import json
from openpyxl import load_workbook
from datetime import datetime, time

def excel_zu_textdatei(datei_pfad, output_pfad):
    # Excel-Datei laden
    wb = load_workbook(datei_pfad)
    sheet = wb.active  # Aktives Arbeitsblatt auswählen

    # Dictionary zum Speichern der Daten
    daten_dict = {}

    # Annahme: Die erste Zeile enthält die Spaltenüberschriften
    spalten = [cell.value for cell in sheet[1]]

    # Durch die Zeilen iterieren und Daten in das Dictionary einfügen
    for row in sheet.iter_rows(min_row=2, values_only=True):
        eintrag = {}
        for i in range(len(spalten)):
            wert = row[i]
            # Überprüfen, ob der Wert ein datetime- oder time-Objekt ist
            if isinstance(wert, datetime):
                eintrag[spalten[i]] = wert.isoformat()  # In ISO-Format umwandeln
            elif isinstance(wert, time):
                eintrag[spalten[i]] = wert.strftime('%H:%M')  # Zeit im HH:MM-Format umwandeln
            else:
                eintrag[spalten[i]] = wert
        zeilen_index = len(daten_dict) + 1  # Zeilenindex als Schlüssel verwenden
        daten_dict[zeilen_index] = eintrag  # Eintrag hinzufügen

    # Daten in eine JSON-Datei speichern
    with open(output_pfad, 'w') as json_file:
        json.dump(daten_dict, json_file, ensure_ascii=False, indent=4)


def text_zu_excel(datei_pfad, input_pfad, spalten_name):
    # Excel-Datei laden
    wb = load_workbook(datei_pfad)
    sheet = wb.active  # Aktives Arbeitsblatt auswählen

    # JSON-Daten laden
    with open(input_pfad, 'r') as json_file:
        neue_daten = json.load(json_file)

    # Spalte aktualisieren
    aktualisiere_spalte(sheet, spalten_name, neue_daten)

    # Änderungen in der Excel-Datei speichern
    wb.save(datei_pfad)

def aktualisiere_spalte(sheet, spalten_name, neue_daten):
    # Überprüfen, ob die Spalte existiert
    spalten = [cell.value for cell in sheet[1]]
    
    if spalten_name in spalten:
        spalten_index = spalten.index(spalten_name) + 1  # +1, da die Indizes in openpyxl 1-basiert sind
        print(f"Aktualisiere Spalte: {spalten_name} (Index: {spalten_index})")
        
        for row_index in range(2, sheet.max_row + 1):
            # Hier wird der neue Wert aus dem Dictionary geholt
            zeilen_index = row_index - 2  # Um den Index im Dictionary zu erhalten
            if zeilen_index < len(neue_daten):
                neue_daten_wert = neue_daten.get(str(zeilen_index + 1))  # +1, da der Index im Dictionary 1-basiert ist
                if (neue_daten_wert != None):
                    # Werte in die Zellen eintragen
                    for spalte, wert in neue_daten_wert.items():
                        if spalte in spalten:  # Überprüfen, ob die Spalte existiert
                            spalte_index = spalten.index(spalte) + 1  # +1 für 1-basierte Indizes
                            sheet.cell(row=row_index, column=spalte_index, value=wert)
                else:
                    print(f"Warnung: Keine Daten für Zeilenindex {zeilen_index + 1} gefunden.")
    else:
        print(f"Spalte '{spalten_name}' nicht gefunden. Füge neue Spalte hinzu.")
        neue_spalte_index = len(spalten) + 1  # Index für die neue Spalte
        sheet.cell(row=1, column=neue_spalte_index, value=spalten_name)  # Überschrift für die neue Spalte
        
        for row_index in range(2, sheet.max_row + 1):
            zeilen_index = row_index - 2  # Um den Index im Dictionary zu erhalten
            if zeilen_index < len(neue_daten):
                neue_daten_wert = neue_daten.get(str(zeilen_index + 1))  # +1, da der Index im Dictionary 1-basiert ist
                if (neue_daten_wert != None):
                    # Werte in die Zellen eintragen
                    for spalte, wert in neue_daten_wert.items():
                        spalte_index = neue_spalte_index  # Neue Spalte verwenden
                        sheet.cell(row=row_index, column=spalte_index, value=wert)
                else:
                    print(f"Warnung: Keine Daten für Zeilenindex {zeilen_index + 1} gefunden.")

              
                
# Beispielaufrufe der Funktionen
excel_zu_textdatei('Daten.xlsx', 'daten_output.txt')
text_zu_excel('Daten.xlsx', 'daten_input.txt', 'Done')
